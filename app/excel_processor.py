"""
Clean Excel image processor for BOM2Pic.
Extracts images from Excel files and creates ZIP downloads.
"""
import io
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import List, NamedTuple

import openpyxl
from PIL import Image


class ImageData(NamedTuple):
    """Container for extracted image data."""
    name_raw: str
    image_bytes: bytes
    sheet: str
    row: int


def column_letter_to_index(letter: str) -> int:
    """Convert Excel column letter (A, B, C...) to zero-based index."""
    letter = letter.upper().strip()
    if not letter.isalpha() or len(letter) > 2:
        raise ValueError(f"Invalid column letter: {letter}")
    
    result = 0
    for char in letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1


def extract_images_from_excel(xlsx_bytes: bytes, image_col_letter: str, name_col_letter: str) -> List[ImageData]:
    """
    Extract images from Excel file.
    
    Args:
        xlsx_bytes: Excel file content as bytes
        image_col_letter: Column containing images (e.g., 'A')
        name_col_letter: Column containing names (e.g., 'B')
        
    Returns:
        List of ImageData objects
    """
    try:
        # Convert column letters to indices
        image_col_idx = column_letter_to_index(image_col_letter)
        name_col_idx = column_letter_to_index(name_col_letter)
        
        # Load workbook
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        extracted_images = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Get images from sheet
            if hasattr(sheet, '_images'):
                for img in sheet._images:
                    # Find the cell containing this image
                    if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                        row_idx = img.anchor._from.row
                        col_idx = img.anchor._from.col
                        
                        # Check if image is in the target column
                        if col_idx == image_col_idx:
                            # Get the name from the name column
                            name_cell = sheet.cell(row=row_idx + 1, column=name_col_idx + 1)
                            name_raw = str(name_cell.value) if name_cell.value else f"image_{row_idx}"
                            
                            # Extract image bytes
                            image_bytes = img.ref.getvalue()
                            
                            extracted_images.append(ImageData(
                                name_raw=name_raw,
                                image_bytes=image_bytes,
                                sheet=sheet_name,
                                row=row_idx + 1
                            ))
        
        return extracted_images
        
    except Exception as e:
        raise ValueError(f"Failed to process Excel file: {str(e)}")


def normalize_filename(name: str) -> str:
    """Convert name to safe filename."""
    if not name or not name.strip():
        return "image"
    
    # Clean the name
    clean = str(name).strip()
    clean = re.sub(r'[<>:"/\\|?*]', '', clean)  # Remove invalid chars
    clean = re.sub(r'\s+', '_', clean)  # Replace spaces with underscores
    clean = clean[:50]  # Limit length
    
    return clean or "image"


def detect_image_extension(image_bytes: bytes) -> str:
    """Detect image format and return appropriate extension."""
    try:
        with Image.open(io.BytesIO(image_bytes)) as img:
            format_map = {
                'JPEG': 'jpg',
                'PNG': 'png',
                'GIF': 'gif',
                'BMP': 'bmp',
                'TIFF': 'tif',
                'WEBP': 'webp'
            }
            return format_map.get(img.format, 'png')
    except:
        return 'png'  # Default fallback


def create_images_zip(images: List[ImageData]) -> io.BytesIO:
    """
    Create a ZIP file containing all images and a report.
    
    Args:
        images: List of ImageData objects
        
    Returns:
        io.BytesIO containing ZIP file
    """
    if not images:
        raise ValueError("No images to process")
    
    zip_buffer = io.BytesIO()
    seen_names = set()
    saved_count = 0
    duplicate_count = 0
    manifest_rows = []
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Create images directory in ZIP
        images_dir = "images/"
        
        for img_data in images:
            # Create safe filename
            base_name = normalize_filename(img_data.name_raw)
            extension = detect_image_extension(img_data.image_bytes)
            filename = f"{base_name}.{extension}"
            
            # Handle duplicates
            if filename in seen_names:
                action = "Duplicate"
                duplicate_count += 1
            else:
                action = "Saved"
                saved_count += 1
                seen_names.add(filename)
            
            # Always write (last one wins for duplicates)
            zf.writestr(f"{images_dir}{filename}", img_data.image_bytes)
            
            # Add to manifest
            manifest_rows.append([
                img_data.sheet,
                img_data.row,
                img_data.name_raw,
                filename,
                action
            ])
        
        # Create report CSV
        import csv
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(["sheet", "row", "part_name", "filename", "action"])
        writer.writerows(manifest_rows)
        zf.writestr("report.csv", csv_buffer.getvalue())
    
    zip_buffer.seek(0)
    return zip_buffer, saved_count, duplicate_count


def process_excel_files(files_data: List[tuple], image_column: str, name_column: str) -> tuple:
    """
    Process multiple Excel files and return ZIP with all images.
    
    Args:
        files_data: List of (filename, file_bytes) tuples
        image_column: Column letter for images
        name_column: Column letter for names
        
    Returns:
        (zip_buffer, total_processed, saved_count, duplicate_count)
    """
    all_images = []
    
    for filename, file_bytes in files_data:
        try:
            images = extract_images_from_excel(file_bytes, image_column, name_column)
            all_images.extend(images)
        except Exception as e:
            # Skip failed files and continue with others
            continue
    
    if not all_images:
        raise ValueError("No images found in any of the uploaded files")
    
    zip_buffer, saved_count, duplicate_count = create_images_zip(all_images)
    
    return zip_buffer, len(all_images), saved_count, duplicate_count
